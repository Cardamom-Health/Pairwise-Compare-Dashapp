import dash  
from dash import dcc, html, Input, Output, State, dash_table  
import pandas as pd  
import io, base64  
import dash_bootstrap_components as dbc  
from openpyxl import Workbook, load_workbook  
from openpyxl.styles import PatternFill, Font  
from itertools import combinations, permutations
  
def parse_contents(contents, filename):  
    content_type, content_string = contents.split(',')  
    decoded = base64.b64decode(content_string)  
    try:  
        if filename.lower().endswith('.csv'):  
            return pd.read_csv(io.StringIO(decoded.decode('utf-8')))  
        elif filename.lower().endswith(('.xls', '.xlsx')):  
            return pd.read_excel(io.BytesIO(decoded))  
    except Exception as e:  
        print(f"Error parsing {filename}: {e}")  
    return pd.DataFrame()  
  
def parse_metadata_string(s):  
    if pd.isnull(s):  
        return set()  
    return set(a.strip() for a in s.split(',') if a.strip())  
  
app = dash.Dash(__name__, suppress_callback_exceptions=True, external_stylesheets=[dbc.themes.LUX])  
server = app.server  
  
app.layout = html.Div([  
    html.H2("Pairwise Record Comparison"),  
    html.Div([  
        html.Button('(Optional) Generate Pairwise Table from List of IDs', id='gen-pairs-btn', n_clicks=0),  
        html.Div(id='upload-list-div', style={'marginTop':10}),  
        dcc.Download(id='download-pairs-table'),  
    ]),  
    html.Hr(style={'marginTop':20}), 
    html.Div([  
        html.Div([  
            html.H5("Upload Pairwise Table (with two IDs per row)"),  
            dcc.Upload(id='upload-pairs',  
                children=html.Button('Upload Pairs Table'), multiple=False),  
            html.Div(id='pairs-uploaded', style={'marginBottom':10, 'color':'green'})  
        ], style={'width':'49%', 'display':'inline-block'}),  
        html.Div([  
            html.H5("Upload ID Metadata Table"),  
            dcc.Upload(id='upload-lookup',  
                children=html.Button('Upload Metadata Table'), multiple=False),  
            html.Div(id='lookup-uploaded', style={'marginBottom':10, 'color':'green'})  
        ], style={'width':'49%', 'display':'inline-block'}),  
    ]),  
    html.Br(),  
    # Dummy dropdowns for suppress_callback_exceptions (hidden)  
    dcc.Dropdown(id='sel-id1', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-id2', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-sim', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-lookup-id', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-lookup-name', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-lookup-usage', options=[], style={'display': 'none'}),  
    dcc.Dropdown(id='sel-lookup-meta', options=[], style={'display': 'none'}),  
    html.Div(id='column-selectors'),  
    html.Br(),  
  
    # Static compare-columns dropdown (always present)  
    html.Div([  
        html.Label("Columns to compare for shared/unique values (use for attributes/metadata):"),  
        dcc.Dropdown(  
            id='compare-columns',  
            options=[],  
            value=[],  
            multi=True  
        ),  
    ], style={'marginBottom': '15px'}),  
  
    html.Br(),  
    html.Div(id="display-column-selector"),  
    html.Button("Show Merged Table", id='show-btn', n_clicks=0, style={'marginTop': '10px'}),  
    html.Hr(),  
    html.Div([  
        html.Button("Export to Excel", id='export-btn', n_clicks=0, style={'marginRight': '10px'}),  
        dcc.Download(id="download-xlsx")  
    ]),  
    html.Br(), 
    # Always-present DataTable  
    dash_table.DataTable(  
        id='main-table',  
        data=[],  
        columns=[],  
        page_size=10,  
        row_selectable='single',  
        sort_action='native',      # Enable sorting  
        filter_action='native',    # Enable filtering  
        selected_rows=[],  
        style_table={'overflowX': 'auto'},  
        style_cell={'minWidth':'120px', 'whiteSpace':'normal'},  
        style_header={'backgroundColor':'#f4f4f4', 'fontWeight':700},  
    ),  
    html.Br(), 
    html.Div(id='comparison-card')  
])  

@app.callback(  
    Output('upload-list-div', 'children'),  
    Input('gen-pairs-btn', 'n_clicks')  
)  
def show_list_upload(n_clicks):  
    if n_clicks == 0:  
        return ""  
    return html.Div([  
        html.H5("Upload list of IDs (CSV or Excel; single column named 'ID' or similar)"),  
        dcc.Upload(id='upload-id-list', children=html.Button('Upload ID List'), multiple=False),  
        html.Div(id='id-list-uploaded', style={'marginBottom':10,'color':'green'}),  
        html.Button('Download All Pairwise Comparisons', id='download-pairs-btn', n_clicks=0, style={'marginTop':10})  
    ])  
  
@app.callback(  
    Output('id-list-uploaded', 'children'),  
    Input('upload-id-list', 'filename')  
)  
def show_id_list_filename(filename):  
    return f"File uploaded: {filename}" if filename else ""  
  
@app.callback(  
    Output('download-pairs-table', 'data'),  
    Input('download-pairs-btn', 'n_clicks'),  
    State('upload-id-list', 'contents'),  
    State('upload-id-list', 'filename'),  
    prevent_initial_call=True  
)  
def make_pairs(n_clicks, contents, filename):  
    if not n_clicks or not contents or not filename:  
        return dash.no_update  
    # Parse the upload as you do elsewhere:  
    df = parse_contents(contents, filename)  
    if df.empty:  
        return dash.no_update  
    # Try to pick the column to use for IDs:  
    id_col = next((c for c in df.columns if 'id' in c.lower()), df.columns[0])  
    ids = df[id_col].dropna().unique().tolist()  
    # -- choose combinations (unordered, no repeats) or permutations (ordered, no repeats) --  
    # Usually for record linkage: use unordered pairs.  
    pair_rows = [{'ID1': a, 'ID2': b} for a, b in combinations(ids, 2)]  
    out_df = pd.DataFrame(pair_rows)  
    # Save as Excel in memory for download  
    output = io.BytesIO()  
    with pd.ExcelWriter(output, engine='openpyxl') as writer:  
        out_df.to_excel(writer, index=False, sheet_name='Pairs')  
    output.seek(0)  
    return dcc.send_bytes(output.getvalue(), 'pairs_table.xlsx') 

@app.callback(  
    Output('pairs-uploaded', 'children'),  
    Output('lookup-uploaded', 'children'),  
    Input('upload-pairs', 'filename'),  
    Input('upload-lookup', 'filename')  
)  
def show_filenames(pairs_name, lookup_name):  
    up1 = f"File uploaded: {pairs_name}" if pairs_name else ""  
    up2 = f"File uploaded: {lookup_name}" if lookup_name else ""  
    return up1, up2  
  
@app.callback(  
    Output('column-selectors', 'children'),  
    Input('upload-pairs', 'contents'),  
    Input('upload-pairs', 'filename'),  
    Input('upload-lookup', 'contents'),  
    Input('upload-lookup', 'filename'),  
)  
def update_column_selectors(pairs_content, pairs_name, lookup_content, lookup_name):  
    if not pairs_content or not lookup_content:  
        return ""  
    pairs_df = parse_contents(pairs_content, pairs_name)  
    lookup_df = parse_contents(lookup_content, lookup_name)  
    if pairs_df.empty or lookup_df.empty:  
        return html.Div("One or both files could not be read. Please re-upload.")  
  
    pair_cols = [{"label": c, "value": c} for c in pairs_df.columns]  
    lookup_cols = [{"label": c, "value": c} for c in lookup_df.columns]  
    usage_cols = [  
        {"label": c, "value": c} for c in lookup_df.select_dtypes(include="number").columns  
    ]  
    meta_cols = [  
        {"label": c, "value": c} for c in lookup_df.select_dtypes(include="object").columns  
    ]  
    def guess(cols, hints):  
        for h in hints:  
            for c in cols:  
                if h.lower() in c.lower():  
                    return c  
        return None  
    id1_guess = guess(pairs_df.columns, ["id1", "query", "id_1", "id 1"])  
    id2_guess = guess(pairs_df.columns, ["id2", "subject", "id_2", "id 2"])  
    if not id1_guess or not id2_guess:  
        ids = [c for c in pairs_df.columns if 'id' in c.lower()]  
        if len(ids) >= 2:  
            id1_guess, id2_guess = ids[0], ids[1]  
    sim_guess = guess(pairs_df.columns, ["sim", "score", "similarity"])  
    lookup_id_guess = guess(lookup_df.columns, ["id"])  
    lookup_name_guess = guess(lookup_df.columns, ["name"])  
    usage_guess = guess(lookup_df.select_dtypes(include="number").columns, ["usage", "count", "amount", "score"])  
    meta_guess = guess(lookup_df.select_dtypes(include="object").columns, ["meta", "attribute", "attr"])  
  
    return html.Div([  
        html.Div([  
            html.Div([  
                html.Label("Pairs table - ID 1 column:"),  
                dcc.Dropdown(id='sel-id1', options=pair_cols, value=id1_guess),  
            ], style={'width': '30%', 'display': 'inline-block', 'marginRight': '5%'}),  
            html.Div([  
                html.Label("Pairs table - ID 2 column:"),  
                dcc.Dropdown(id='sel-id2', options=pair_cols, value=id2_guess),  
            ], style={'width': '30%', 'display': 'inline-block', 'marginRight': '5%'}),  
            html.Div([  
                html.Label("(Optional) Similarity/Score column:"),  
                dcc.Dropdown(id='sel-sim', options=pair_cols, value=sim_guess, clearable=True),  
            ], style={'width': '30%', 'display': 'inline-block'}),  
        ], style={'marginBottom': '15px'}),  
        html.Div([  
            html.Div([  
                html.Label("Details table - ID column:"),  
                dcc.Dropdown(id='sel-lookup-id', options=lookup_cols, value=lookup_id_guess),  
            ], style={'width': '30%', 'display': 'inline-block', 'marginRight': '5%'}),  
            html.Div([  
                html.Label("Details table - Name column:"),  
                dcc.Dropdown(id='sel-lookup-name', options=lookup_cols, value=lookup_name_guess),  
            ], style={'width': '30%', 'display': 'inline-block', 'marginRight': '5%'}),  
            html.Div([  
                html.Label("Details table - Usage column:"),  
                dcc.Dropdown(id='sel-lookup-usage', options=usage_cols, value=usage_guess),  
            ], style={'width': '15%', 'display': 'inline-block', 'marginRight': '5%'}),  
            html.Div([  
                html.Label("Details table - Metadata column:"),  
                dcc.Dropdown(id='sel-lookup-meta', options=meta_cols, value=meta_guess),  
            ], style={'width': '20%', 'display': 'inline-block'}),  
        ]),  
    ])  
  
@app.callback(  
    Output('compare-columns', 'options'),  
    Output('compare-columns', 'value'),  
    Input('upload-lookup', 'contents'),  
    Input('upload-lookup', 'filename'),  
    Input('sel-lookup-id', 'value'),  
    Input('sel-lookup-meta', 'value'),  
    prevent_initial_call=True  
)  
def update_compare_columns_dropdown(lookup_content, lookup_name, sel_id, sel_meta):  
    if not lookup_content or not lookup_name or not sel_id:  
        return [], []  
    lookup_df = parse_contents(lookup_content, lookup_name)  
    if lookup_df.empty:  
        return [], []  
    all_cols = lookup_df.columns  
    options = [{"label": c, "value": c} for c in all_cols]    
    default_val = [sel_meta] if sel_meta in all_cols else []  
    return options, default_val  
  
@app.callback(  
    Output("display-column-selector", "children"),  
    Input('column-selectors', 'children'),  
    State('sel-id1', 'value'),  
    State('sel-id2', 'value'),  
    State('sel-sim', 'value'),  
    State('sel-lookup-name', 'value'),  
    State('sel-lookup-usage', 'value'),  
)  
def update_display_column_selector(_, id1_col, id2_col, sim_col, name_col, usage_col):  
    if not (id1_col and id2_col):  
        return ""  
    options = [  
        {"label": "ID 1", "value": "ID_1"},  
        {"label": "ID 2", "value": "ID_2"},  
    ]  
    if name_col:  
        options += [  
            {"label": "Name 1", "value": "Name_1"},  
            {"label": "Name 2", "value": "Name_2"},  
        ]  
    if usage_col:  
        options += [  
            {"label": f"{usage_col} (ID 1)", "value": f"{usage_col}_1"},  
            {"label": f"{usage_col} (ID 2)", "value": f"{usage_col}_2"},  
        ]  
    if sim_col:  
        options.append({"label": "Similarity/Score", "value": "Similarity/Score"})  
    default_value = [o["value"] for o in options]  
    return html.Div([  
        html.Label("Columns to display for ID info:"),  
        dcc.Dropdown(  
            id="display-columns",  
            options=options,  
            value=default_value,  
            multi=True  
        ),  
        html.Br(),  
    ])  
  
@app.callback(  
    Output("download-xlsx", "data"),  
    Input("export-btn", "n_clicks"),  
    State('main-table', 'data'),  
    State('main-table', 'columns'),  
    State('compare-columns', 'value'),  
    prevent_initial_call=True  
)  
def export_to_excel(n_clicks, data, columns, compare_cols):  
    if not n_clicks or not data or not columns:  
        return dash.no_update  
  
    # Step 1: Build DataFrame and save to Excel in memory with openpyxl  
    df = pd.DataFrame(data)  
    output = io.BytesIO()  
    with pd.ExcelWriter(output, engine="openpyxl") as writer:  
        df.to_excel(writer, index=False, sheet_name="Merged")  
    output.seek(0)  
    wb = load_workbook(output)  
    ws = wb["Merged"]  
  
    # Step 2: Bold headers  
    header_font = Font(bold=True)  
    for cell in ws[1]:  
        cell.font = header_font  
  
    # Step 3: Adjust column widths automatically  
    for col in ws.columns:  
        max_length = 0  
        col_letter = col[0].column_letter  
        for cell in col:  
            try:  
                value = str(cell.value)  
            except:  
                value = ""  
            if value:  
                max_length = max(max_length, len(value))  
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)  # leave a little padding  
  
    # Step 4: Color relevant columns  
    color_map = {}  
    if compare_cols:  
        for col in compare_cols:  
            color_map[f"{col} | Shared in both"] = "D6F5D6"  
            color_map[f"{col} | Unique to ID 1"] = "FFFACD"  
            color_map[f"{col} | Unique to ID 2"] = "FFD9EC"  
  
    # Openpyxl rows/columns are 1-indexed and include header row  
    col_indices = {cell.value: cell.column for cell in ws[1]}  
  
    for col, fillcolor in color_map.items():  
        if col not in col_indices:  
            continue  
        col_idx = col_indices[col]  
        fill = PatternFill(start_color=fillcolor, end_color=fillcolor, fill_type="solid")  
        # start from row 2 (data)  
        for row in range(2, ws.max_row + 1):  
            cell = ws.cell(row=row, column=col_idx)  
            if cell.value is not None and str(cell.value).strip() != "":  
                cell.fill = fill  
  
    # Save workbook to BytesIO for download  
    final_output = io.BytesIO()  
    wb.save(final_output)  
    final_output.seek(0)  
    return dcc.send_bytes(final_output.getvalue(), "merged_comparison.xlsx") 
 
@app.callback(  
    Output('main-table', 'data'),  
    Output('main-table', 'columns'),  
    Output('main-table', 'style_data_conditional'),  
    Output('main-table', 'selected_rows'),  
    Input('show-btn', 'n_clicks'),  
    State('sel-id1', 'value'),  
    State('sel-id2', 'value'),  
    State('sel-sim', 'value'),  
    State('sel-lookup-id', 'value'),  
    State('sel-lookup-name', 'value'),  
    State('sel-lookup-usage', 'value'),  
    State('sel-lookup-meta', 'value'),  
    State('compare-columns', 'value'),  
    State('display-columns', 'value'),  
    State('upload-pairs', 'contents'),  
    State('upload-pairs', 'filename'),  
    State('upload-lookup', 'contents'),  
    State('upload-lookup', 'filename'),  
    prevent_initial_call=True  
)  
def build_main_table(n_clicks, id1_col, id2_col, sim_col, lookup_id_col, name_col, usage_col, meta_col, compare_cols, display_cols, pairs_content, pairs_name, lookup_content, lookup_name):  
    if not pairs_content or not lookup_content:  
        return [], [], [], []  
    pairs_df = parse_contents(pairs_content, pairs_name)  
    lookup_df = parse_contents(lookup_content, lookup_name)  
    if pairs_df.empty or lookup_df.empty:  
        return [], [], [], []  
    merged = pairs_df.copy()  
    merged = merged.rename(columns={id1_col: "ID_1", id2_col: "ID_2"})  
    merged = pd.merge(merged, lookup_df.rename(  
        columns={lookup_id_col: "ID_1", name_col: "Name_1"}  
    )[[ "ID_1", "Name_1"]], on="ID_1", how="left")  
    merged = pd.merge(merged, lookup_df.rename(  
        columns={lookup_id_col: "ID_2", name_col: "Name_2"}  
    )[[ "ID_2", "Name_2"]], on="ID_2", how="left")  
    if usage_col and usage_col in lookup_df.columns:  
        merged = pd.merge(  
            merged,  
            lookup_df[[lookup_id_col, usage_col]].rename(columns={lookup_id_col: "ID_1", usage_col: f"{usage_col}_1"}),  
            on="ID_1",  
            how="left"  
        )  
        merged = pd.merge(  
            merged,  
            lookup_df[[lookup_id_col, usage_col]].rename(columns={lookup_id_col: "ID_2", usage_col: f"{usage_col}_2"}),  
            on="ID_2",  
            how="left"  
        )  
    if sim_col and sim_col in pairs_df.columns:  
        merged["Similarity/Score"] = merged[sim_col]  
        try:  
            merged["Similarity/Score"] = merged["Similarity/Score"].apply(lambda x: round(float(x), 3) if pd.notnull(x) else "")  
        except:  
            pass  
    style_data_conditional = []  
    if compare_cols:  
        for col in compare_cols:  
            if col not in lookup_df.columns:  
                continue  
            col_is_numeric = pd.api.types.is_numeric_dtype(lookup_df[col])  # <<<< MODIFIED 
            id_to_attr = dict(zip(lookup_df[lookup_id_col], lookup_df[col]))  
            shared_list, uniq1_list, uniq2_list = [], [], []  
            for idx, row in merged.iterrows():  
                id1 = row["ID_1"]  
                id2 = row["ID_2"]  
                meta1 = id_to_attr.get(id1, "")  
                meta2 = id_to_attr.get(id2, "")  
                if col_is_numeric:  # <<<< MODIFIED  
                    set1 = set([str(meta1)]) if pd.notnull(meta1) and meta1 != "" else set()  
                    set2 = set([str(meta2)]) if pd.notnull(meta2) and meta2 != "" else set()  
                else:  
                    set1 = parse_metadata_string(meta1)  
                    set2 = parse_metadata_string(meta2)  
                shared = sorted(set1 & set2)  
                uniq1 = sorted(set1 - set2)  
                uniq2 = sorted(set2 - set1)  
                shared_list.append(", ".join(shared))  
                uniq1_list.append(", ".join(uniq1))  
                uniq2_list.append(", ".join(uniq2))  
            col_shared = f"{col} | Shared in both"  
            col_uniq1 = f"{col} | Unique to ID 1"  
            col_uniq2 = f"{col} | Unique to ID 2"  
            merged[col_shared] = shared_list  
            merged[col_uniq1] = uniq1_list  
            merged[col_uniq2] = uniq2_list  
            # Add highlighting for these columns  
            style_data_conditional += [  
                {  
                    "if": {"column_id": col_shared, "filter_query": f'{{{col_shared}}} != ""'},  
                    "backgroundColor": "#D6F5D6", "color": "black"  
                },  
                {  
                    "if": {"column_id": col_uniq1, "filter_query": f'{{{col_uniq1}}} != ""'},  
                    "backgroundColor": "#FFFACD", "color": "black"  
                },  
                {  
                    "if": {"column_id": col_uniq2, "filter_query": f'{{{col_uniq2}}} != ""'},  
                    "backgroundColor": "#FFD9EC", "color": "black"  
                },  
            ]  
    all_compare_cols = []  
    if compare_cols:  
        for col in compare_cols:  
            all_compare_cols += [  
                f"{col} | Shared in both",  
                f"{col} | Unique to ID 1",  
                f"{col} | Unique to ID 2",  
            ]  
    final_cols = [col for col in display_cols if col in merged.columns]  
    if all_compare_cols:  
        final_cols += [col for col in all_compare_cols if col in merged.columns and col not in final_cols]  
    columns = []  
    for col in final_cols:  
        # Decide if this column should be shown as numeric  
        # Get column values from merged DataFrame  
        if pd.api.types.is_numeric_dtype(merged[col]):  
            columns.append({"name": col, "id": col, "type": "numeric"})  
        else:  
            columns.append({"name": col, "id": col, "type": "text"})    
    data = merged[final_cols].to_dict('records')  
    return data, columns, style_data_conditional, []  
  
@app.callback(  
    Output('comparison-card', 'children'),  
    Input('main-table', 'selected_rows'),  
    State('main-table', 'data'),  
    State('compare-columns', 'value')  
)  
def display_similarity(selected_rows, table_data, compare_cols):  
    if not selected_rows or not table_data or not compare_cols:  
        return ""  
    df = pd.DataFrame(table_data)  
    row = df.iloc[selected_rows[0]]  
    cards = []  
    name1 = row.get('Name_1', row.get('ID_1', ''))  
    name2 = row.get('Name_2', row.get('ID_2', ''))  
    usage_1 = None  
    usage_2 = None  
    for col in df.columns:  
        if col.endswith('_1') and isinstance(row[col], (int, float)):  
            usage_1 = row[col]  
        if col.endswith('_2') and isinstance(row[col], (int, float)):  
            usage_2 = row[col]  
    cards.append(html.H3(f"Comparing: {name1} and {name2}"))  
    if usage_1 is not None:  
        cards.append(html.P(f"Usage for {name1}: {usage_1}"))  
    if usage_2 is not None:  
        cards.append(html.P(f"Usage for {name2}: {usage_2}"))  
    for col in compare_cols:  
        shared_col = f"{col} | Shared in both"  
        unique1_col = f"{col} | Unique to ID 1"  
        unique2_col = f"{col} | Unique to ID 2"  
        shared = row.get(shared_col, "")  
        unique_1 = row.get(unique1_col, "")  
        unique_2 = row.get(unique2_col, "")  
    
        def to_list(s):  
            if pd.isnull(s) or not s: return []  
            items = []  
            for part in s.split(';'):  
                items.extend(x.strip() for x in part.split(','))  
            return [x for x in items if x]  
        
        shared_list = to_list(shared)  
        uniq1_list = to_list(unique_1)  
        uniq2_list = to_list(unique_2)  
        
        # >>> Here is where we add counts!  
        cards.append(html.Div([  
            html.H4(f"Comparison for '{col}'"),  
            html.Div([  
                html.Div([  
                    html.H5(f"Shared (n = {len(shared_list)})"),  
                    html.Ul([html.Li(x) for x in shared_list] or ["None"])  
                ], style={'width': '30%', 'display': 'inline-block', 'verticalAlign': 'top', 'backgroundColor': '#D6F5D6', 'padding': '10px', 'margin': '10px'}),  
                html.Div([  
                    html.H5(f"Unique to {name1} (n = {len(uniq1_list)})"),  
                    html.Ul([html.Li(x) for x in uniq1_list] or ["None"])  
                ], style={'width': '30%', 'display': 'inline-block', 'verticalAlign': 'top', 'backgroundColor': '#FFFACD', 'padding': '10px', 'margin': '10px'}),  
                html.Div([  
                    html.H5(f"Unique to {name2} (n = {len(uniq2_list)})"),  
                    html.Ul([html.Li(x) for x in uniq2_list] or ["None"])  
                ], style={'width': '30%', 'display': 'inline-block', 'verticalAlign': 'top', 'backgroundColor': '#FFD9EC', 'padding': '10px', 'margin': '10px'}),  
            ], style={'display': 'flex', 'justifyContent': 'space-around'})  
        ], style={'marginTop': 30, 'marginBottom': 30}))  
    return html.Div(cards)  
  
if __name__ == '__main__':  
    app.run_server(debug=True)  